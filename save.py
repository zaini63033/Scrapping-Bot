import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from standard import initialize_driver, parse_record

def save_to_spreadsheet(county_name, results, save_path):
    # Define the path to the spreadsheet
    file_path = os.path.join(save_path, "output.xlsx")

    # Check if the spreadsheet already exists
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()

    # Remove the default sheet if it exists
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # Create a new sheet for the county or select the existing one
    if county_name in workbook.sheetnames:
        sheet = workbook[county_name]
    else:
        sheet = workbook.create_sheet(title=county_name)

    # Define the yellow fill for header cells and grey fill for separator row
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Write headers if the sheet is new
    if sheet.max_row == 1:
        headers = [
            "PT-61", "Sale Date", "Actual Price", 
            "Owner First Name", "Owner Middle Name", "Owner Last Name", 
            "Owner Address", "Owner City", "Owner State", "Owner ZIP Code", 
            "Executor First Name", "Executor Middle Name", "Executor Last Name", 
            "Executor Address", "Executor City", "Executor State", "Executor ZIP Code", 
            "Property Address", "Property City", "Property State", "Property ZIP Code", 
            "Map & Parcel Number"
        ]
        sheet.append(headers)

        # Bold the headers and set background color
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = yellow_fill
        sheet.row_dimensions[1].height = 30  # Increase the height of the header row

        # Create a thin separator row
        separator_row_number = 2
        sheet.append([None] * len(headers))  # Add an empty row for the separator
        separator_row = sheet[separator_row_number]
        
        # Merge all cells in the separator row
        sheet.merge_cells(start_row=separator_row_number, start_column=1, end_row=separator_row_number, end_column=len(headers))
        
        # Set the fill color and minimal height for the separator row
        for cell in separator_row:
            cell.fill = grey_fill
        sheet.row_dimensions[separator_row_number].height = 5  # Minimal height for the separator

    # Initialize the Selenium WebDriver
    driver = initialize_driver()

    # Append each result to the sheet
    for record in results:
        parsed_record = parse_record(record, driver)
        sheet.append([
            parsed_record["pt61"],
            parsed_record["sale_date"],
            parsed_record["actual_price"],
            parsed_record["owner_name"]['first_name'],
            parsed_record["owner_name"]['middle_name'],
            parsed_record["owner_name"]['last_name'],
            parsed_record["owner_address"]['address'],
            parsed_record["owner_address"]['city'],
            parsed_record["owner_address"]['state'],
            parsed_record["owner_address"]['zip'],
            parsed_record["executor_name"]['first_name'],
            parsed_record["executor_name"]['middle_name'],
            parsed_record["executor_name"]['last_name'],
            parsed_record["executor_address"]['address'],
            parsed_record["executor_address"]['city'],
            parsed_record["executor_address"]['state'],
            parsed_record["executor_address"]['zip'],
            parsed_record["property_address"]['address'],
            parsed_record["property_address"]['city'],
            parsed_record["property_address"]['state'],
            parsed_record["property_address"]['zip'],
            parsed_record["map_parcel_number"]
        ])

    # Adjust column widths to fit content
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Get the column name
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 5)  # Add padding for better view
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook and close the driver
    workbook.save(file_path)
    driver.quit()
    workbook.close()
